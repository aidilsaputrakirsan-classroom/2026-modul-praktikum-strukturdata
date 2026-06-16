"""
Compute final grades for Kelas A & B Struktur Data.

Komposisi bobot:
- Praktikum: 30%  (rata-rata M1-M12 dari README evaluasi)
- Kuis:      10%
- UTS:       25%
- UAS:       25%  (belum ada - kolom placeholder)
- Kehadiran: 10%  (HADIR=1, IZIN=0.5, TANPA KETERANGAN=0, null=skip)

Konversi huruf (batas bawah inklusif):
- A:  >= 86
- AB: 76 <= x < 86
- B:  66 <= x < 76
- BC: 56 <= x < 66
- C:  51 <= x < 56
- D:  41 <= x < 51
- E:  < 41

Catatan:
- Daftar mahasiswa diambil dari file Excel (presensi), sehingga setiap NIM
  yang terdaftar tetap masuk laporan meskipun tidak ada data evaluasi.
- Sel kosong di tabel menggunakan tanda '-'.
- Kolom Huruf hanya menampilkan huruf (tanpa badge skor).
"""

import os
import openpyxl
from pathlib import Path

# File presensi memuat Kuis/UTS/Kehadiran. File ini sengaja TIDAK disimpan di
# repo (privasi). Set lokasi via env var PRESENSI_XLSX saat menjalankan script,
# contoh: PRESENSI_XLSX=/path/ke/presensi.xlsx python compute_final_grades.py
EXCEL_PATH = Path(os.environ.get('PRESENSI_XLSX', 'PRESENSI PRAKTIKUM STRUKTUR DATA.xlsx'))

BOBOT = {
    'praktikum': 0.30,
    'kuis': 0.10,
    'uts': 0.25,
    'uas': 0.25,
    'hadir': 0.10,
}


def letter_grade(score):
    if score is None:
        return '-'
    if score >= 86:
        return 'A'
    if score >= 76:
        return 'AB'
    if score >= 66:
        return 'B'
    if score >= 56:
        return 'BC'
    if score >= 51:
        return 'C'
    if score >= 41:
        return 'D'
    return 'E'


def safe_num(v):
    if v is None or v == '' or (isinstance(v, str) and v.strip() == ''):
        return None
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


def compute_attendance(row_values, week_active):
    """Hitung kehadiran. `week_active` adalah list bool sepanjang row_values
    yang menandai minggu mana yang sudah berjalan (ada minimal 1 HADIR di kelas).

    - HADIR -> 1.0
    - IZIN / SAKIT -> 0.5
    - TANPA KETERANGAN / ALPA / sel kosong di minggu yang SUDAH berjalan -> 0
    - sel kosong di minggu yang BELUM berjalan -> skip
    """
    weight_sum = 0.0
    max_sum = 0.0
    for i, v in enumerate(row_values):
        is_empty = v is None or (isinstance(v, str) and v.strip() == '')
        if is_empty:
            if i < len(week_active) and week_active[i]:
                max_sum += 1.0  # minggu sudah berjalan, sel kosong = alpa
            continue
        s = str(v).strip().upper()
        if s == 'HADIR':
            weight_sum += 1.0
            max_sum += 1.0
        elif s in ('IZIN', 'SAKIT'):
            weight_sum += 0.5
            max_sum += 1.0
        else:  # TANPA KETERANGAN, ALPA, atau lainnya
            max_sum += 1.0
    if max_sum == 0:
        return None
    return (weight_sum / max_sum) * 100.0


def detect_active_weeks(sheet, n_cols):
    """Return list bool: True jika minggu (kolom ke-i) sudah berjalan.
    Minggu dianggap sudah berjalan jika ada >= 1 entry HADIR/IZIN di kolom tsb.
    """
    active = [False] * n_cols
    for row in sheet.iter_rows(min_row=3, values_only=True):
        cells = row[2:2 + n_cols]
        for i, v in enumerate(cells):
            if v is None or (isinstance(v, str) and v.strip() == ''):
                continue
            s = str(v).strip().upper()
            if s in ('HADIR', 'IZIN', 'SAKIT', 'TANPA KETERANGAN', 'ALPA'):
                active[i] = True
    return active


def parse_excel():
    """Return: {kelas: {nim: {nama_excel, kuis, uts, attendance}}} for ALL NIMs in Excel."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    data = {'A': {}, 'B': {}}

    for kelas, sheet in [('A', 'Kuis dan UTS KELAS A'), ('B', 'Kuis dan UTS KELAS B')]:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            nim = row[0]
            if not nim:
                continue
            try:
                nim = int(nim)
            except (ValueError, TypeError):
                continue
            data[kelas][nim] = {
                'nama_excel': row[1],
                'kuis': safe_num(row[2]),
                'uts': safe_num(row[3]),
                'attendance': None,
            }

    for kelas, sheet in [('A', 'KELAS A'), ('B', 'KELAS B')]:
        ws = wb[sheet]
        n_cols = ws.max_column - 2  # kolom A=NIM, B=Nama, C+ = minggu
        active_weeks = detect_active_weeks(ws, n_cols)
        for row in ws.iter_rows(min_row=3, values_only=True):
            nim = row[0]
            if not nim:
                continue
            try:
                nim = int(nim)
            except (ValueError, TypeError):
                continue
            att = compute_attendance(row[2:2 + n_cols], active_weeks)
            if nim not in data[kelas]:
                data[kelas][nim] = {
                    'nama_excel': row[1],
                    'kuis': None,
                    'uts': None,
                    'attendance': att,
                }
            else:
                data[kelas][nim]['attendance'] = att

    return data


# Praktikum scores per NIM (dari README evaluasi). NIM yang tidak ada di sini
# berarti tidak ada data evaluasi praktikum (akan ditampilkan '-').
PRAKTIKUM_A = {  # rata-rata M1-7, 9-15 (14 modul); update M15 Trie/Tensor
    10231081: 91.64,  # M14=83 (T1 blok analisis tidak diisi)
    10251002: 95.29,
    10251005: 56.71,  # M14=0 (tidak ada Modul 14)
    10251008: 62.50,  # M14=0 (hanya praktikum)
    10251011: 77.64,  # M14=33 (hanya T2)
    10251014: 89.93,
    10251017: 84.93,
    10251020: 99.29,
    10251023: 90.00,  # koreksi feedback 13Jun
    10251026: 95.79,  # koreksi 16Jun: M12 dinilai ulang 0->100 (folder minggu12 tak terdeteksi saat evaluasi awal)
    10251029: 90.57,
    10251032: 92.79,  # koreksi feedback 13Jun
    10251035: 97.43,
    10251038: 96.00,
    10251044: 84.57,  # koreksi feedback 13Jun
    10251047: 83.71,  # M14=0 (hanya praktikum)
    10251050: 81.00,  # M14=0 (tidak ada Modul 14)
    10251053: 86.93,  # koreksi feedback 13Jun
    10251056: 92.14,
    10251059: 85.71,  # koreksi feedback 13Jun
    10251062: 95.29,  # koreksi feedback 13Jun
    10251065: 97.64,  # koreksi feedback 13Jun
    10251068: 87.86,
    10251071: 96.43,  # koreksi feedback 13Jun
    10251074: 88.14,  # koreksi feedback 13Jun
    10251086: 90.50,  # Hidayah pakai NIM 10251086 di Excel (terdaftar GitHub Classroom: 10251075); M14=100
    10251077: 80.71,  # koreksi feedback 13Jun
    10251080: 85.71,  # koreksi feedback 13Jun
    10251083: 97.64,
    10251089: 86.93,  # koreksi feedback 13Jun
    10251092: 77.36,
    10251095: 96.43,
    10251104: 96.21,  # koreksi feedback 13Jun
    10251107: 75.29,  # M14=0 (hanya praktikum)
    10251110: 93.50,
    10251114: 77.43,
    10251117: 88.29,  # M14=83 (T1 blok analisis tidak diisi)
}

PRAKTIKUM_B = {  # rata-rata M1-7, 9-15 (14 modul); update M15 Trie/Tensor
    10231001: 97.64,
    10231038: 91.14,
    10251006: 88.36,
    10251009: 42.86,  # M14=0
    10251012: 83.36,  # koreksi: M9 & M10 ada (folder Minggu9/Minggu10 tak terdeteksi) - M9=100, M10=67
    10251015: 76.21,
    10251018: 89.07,
    10251021: 55.71,  # koreksi: M13 ada (terblokir path invalid `...`) - M13=100
    10251024: 85.71,  # M14=0
    10251027: 81.00,  # M14=0 (placeholder kosong)
    10251084: 71.71,  # M14=0 (hanya praktikum)
    10251033: 72.64,
    10251036: 59.07,  # M14=0
    10251039: 35.71,  # M14=0
    10251042: 33.36,  # M14=0
    10251048: 90.50,
    10251051: 21.43,  # M14=0
    10251054: 68.00,  # M14=0
    10251057: 88.14,  # koreksi: M10 ada (terblokir path invalid `H. L.`) - M10=100
    10251060: 95.29,
    10251063: 90.50,
    10251066: 79.79,
    10251069: 55.00,  # M14=0
    10251072: 92.93,
    10251075: 19.07,  # M14=67 (T1 duplikat konten T2)
    10251078: 92.64,
    10251081: 30.00,  # M14=0
    10251087: 51.21,  # M14=0
    10251090: 89.29,  # M14=0 (konten Modul 15, bukan sorting)
    10251093: 31.00,  # M14=0
    10251099: 83.36,  # M14=0
    10251102: 14.29,  # M14=0
    10251105: 82.43,
    10251108: 65.50,  # M14=0 (hanya praktikum)
    10251112: 56.00,
    10251115: 86.93,
}

# Username (GitHub) per NIM
USERNAME_A = {
    10231081: "NorEndGate", 10251002: "10251002", 10251005: "rifatadlyjuliandra",
    10251008: "nadyaazzahra", 10251011: "Asuci-Maharani", 10251014: "Zacky10251014",
    10251017: "devigustiani", 10251020: "Dhevrant", 10251023: "AryaCahyaNugraha22",
    10251026: "Salsabilaayees", 10251029: "SilvesterJevon", 10251032: "FiryalFarah",
    10251035: "10251035-Putri", 10251038: "10251038", 10251044: "Asrul-6",
    10251047: "syahira10251047", 10251050: "Zennin2511", 10251053: "N3CH0",
    10251056: "SyahrinaAliyyuTunggadewi", 10251059: "10251059-coder",
    10251062: "AnandaEurico", 10251065: "Bryan1065", 10251068: "Fajri-ITK",
    10251071: "MuhRasyaAlKhalifi", 10251074: "MuhammadRizqiAkbar33",
    10251086: "MuhammadRizqiHidayah", 10251077: "RisnaAndriani-10251077",
    10251080: "Julian-Zaky-Saputra", 10251083: "MossesRestuN", 10251089: "10251089-Ryan",
    10251092: "snowyy26", 10251095: "NurHayyu-10251095", 10251104: "Chelsea-10251104",
    10251107: "RifatSuthaAbdillah", 10251110: "10251110", 10251114: "Andre10251114",
    10251117: "azrilalfadhil-svg",
}

USERNAME_B = {
    10231001: "AchmadLyraa", 10231038: "firnifziah", 10251006: "Ridz010",
    10251009: "Taufiqerikhrl", 10251012: "fransiskusezekiel", 10251015: "enjelinhartini",
    10251018: "Faiz-codepy", 10251021: "alejandroaprillio", 10251024: "anawithcode",
    10251027: "kurniala", 10251084: "Vedomzz", 10251033: "MlhamNF",
    10251036: "MarineTaranda", 10251039: "strukturdata-b-10251039", 10251042: "heldaa127",
    10251048: "10251048-boop", 10251051: "AhdyarRamadhan", 10251054: "Rizi-54",
    10251057: "fasya25", 10251060: "Nastai1n-bit", 10251063: "liandizgo-design",
    10251066: "noelrandi7", 10251069: "Binngian", 10251072: "reihan-agil-wijaya",
    10251075: "KatonHillal", 10251078: "naafiannisa", 10251081: "10251081-blip",
    10251087: "nabilfdl", 10251090: "nurasyifaahmad", 10251093: "10251093",
    10251099: "Raihanp-dya", 10251102: "devansgt", 10251105: "10251105-haikal",
    10251108: "10251108-create", 10251112: "galeryc-afk", 10251115: "JonatanBagusKristiawan",
}


def compute_nilai_sementara(prak, kuis, uts, hadir):
    """Nilai sementara (tanpa UAS), dinormalisasi ke skala 100.
    Return None jika SEMUA komponen kosong.
    """
    if prak is None and kuis is None and uts is None and hadir is None:
        return None
    p = prak or 0
    k = kuis or 0
    u = uts or 0
    h = hadir or 0
    total = (p * BOBOT['praktikum'] + k * BOBOT['kuis']
             + u * BOBOT['uts'] + h * BOBOT['hadir'])
    bobot_terisi = (BOBOT['praktikum'] + BOBOT['kuis']
                    + BOBOT['uts'] + BOBOT['hadir'])
    return total / bobot_terisi


def compute_nilai_akhir(prak, kuis, uts, uas, hadir):
    """Nilai akhir dengan formula penuh (termasuk UAS), skala 100.
    Komponen kosong dihitung sebagai 0 (penalti). Return None jika SEMUA kosong.
    """
    if (prak is None and kuis is None and uts is None
            and uas is None and hadir is None):
        return None
    p = prak or 0
    k = kuis or 0
    u = uts or 0
    a = uas or 0
    h = hadir or 0
    return (p * BOBOT['praktikum'] + k * BOBOT['kuis'] + u * BOBOT['uts']
            + a * BOBOT['uas'] + h * BOBOT['hadir'])


def fmt(v, dec=0):
    """Format ke integer (dibulatkan) by default. Pakai dec > 0 untuk desimal."""
    if v is None:
        return '-'
    if dec == 0:
        return str(int(round(v)))
    return f"{v:.{dec}f}"


# Override attendance untuk mahasiswa nonaktif (sesuai data resmi presensi).
# Script auto-detect "minggu sudah berjalan" bisa keliru untuk mhs yang hanya
# hadir 1-2 kali; override di sini memastikan nilainya 0 sesuai input dosen.
ATTENDANCE_OVERRIDE = {
    'A': {
        10251041: 0.0,  # Gusti Aulia
        10251098: 0.0,  # Panji Dwi
        10251101: 0.0,  # Fawwaz
    },
    'B': {
        # Dispensasi penuh (1.0) untuk minggu terdampak sakit/kecelakaan (surat/keterangan resmi):
        10251009: 90.91,  # Taufiq: sakit, dispensasi W9-12 (W4 alpa tetap)
        10251102: 81.82,  # Devan: kecelakaan 3-18 Mei, dispensasi W9 & W10
    },
}


# Kuis override (untuk mhs yang data Kuis-nya kosong/salah di Excel)
KUIS_OVERRIDE = {
    'A': {},
    'B': {
        10251115: 60,  # Jonatan Bagus Kristiawan (Excel kosong, diinput susulan)
        # Tidak ikut kuis -> diganti nilai tugas minggu tertentu, DICAP MAKS 70:
        10251021: 70,  # Alejandro: pengganti = tugas Week 10 (nilai 100, cap 70)
        10251102: 70,  # Devan: pengganti = tugas Week 4 (nilai 100, cap 70)
        10251069: 70,  # Albin: pengganti = tugas Week 6 (nilai 100, cap 70)
    },
}


# UTS Final setelah remedial. Formula:
# UTS_final = MAX(UTS_asli, MIN(UTS_remedial, 65))
# Hanya berisi mhs yang skornya naik karena remedial. Yang lain pakai UTS asli dari Excel.
UTS_FINAL_OVERRIDE = {
    'A': {
        10251107: 65,  # Rifat Sutha (asli 47) - remedial PASSED, sblmnya tak terdeteksi
        10251117: 65,  # Azriel (asli 46) - remedial PASSED
        10251114: 65,  # Andre (asli 43) - remedial PASSED
        10251002: 65,  # Dina (asli 30, remedial cap 65)
        10251005: 65,  # Rifat A (asli 46)
        10251017: 65,  # Devi (asli 45)
        10251023: 65,  # Arya (asli 35)
        10251029: 65,  # Jevon (asli 56)
        10251032: 65,  # Firyal (asli 43)
        10251044: 65,  # Asrul (asli 56)
        10251047: 65,  # Syahira (asli 65) - sama
        10251053: 65,  # Noel (asli 50)
        10251056: 65,  # Syahrina (asli 47)
        10251059: 65,  # Taufiiqul (asli 43)
        10251068: 65,  # Fajri (asli 37)
        10251074: 65,  # Akbar (asli 54)
        10251092: 65,  # Razi (asli 35)
        10251104: 65,  # Chelsea (asli 45)
        10251110: 65,  # Farel (asli 37)
        # Yang remedial tapi UTS asli > remedial: Suci (62 vs 45), Ayudya (53 vs 50), Zennin (55 vs 0)
        # → tidak di-override, tetap pakai UTS asli
    },
    'B': {
        10251021: 65,  # Alejandro (asli 36) - remedial terdeteksi, Soal1/2/3 PASSED
        10251012: 65,  # Fransiskus (asli 51)
        10251027: 65,  # Kurnia (asli 55)
        10251039: 65,  # Khoiry (asli 52)
        10251048: 65,  # Dewi (asli 61)
        10251054: 65,  # Rizky (asli 49)
        10251066: 65,  # Noel R (asli 52)
        10251075: 65,  # Katon (asli 32)
        10251102: 65,  # Devan (asli 28)
        10251105: 65,  # Haikal (asli 26)
        10251108: 65,  # Kevin (asli 26)
        10251112: 65,  # Cindy (asli 62)
        10251036: 65,  # Marine (asli 57) - logika S1/2/3 BENAR (PASSED setelah bypass teks markdown di source); dinilai utuh per kebijakan dosen
        # Yang remedial tapi UTS asli > remedial: Firni (51 vs 45), Enjelin (44 vs 35),
        # Shafwat (59 vs 55), Albin (39 vs 10) → tetap pakai UTS asli
    },
}


# Nilai UAS ASLI (dari presensi). Fernando (10251081) & Gilang (10251096) tidak
# hadir UAS → 0. Mahasiswa nonaktif (Gusti/Panji/Fawwaz/Ari/Anisa) tidak ada nilai.
UAS_ASLI_A = {
    10231081: 36, 10251002: 51, 10251005: 41, 10251008: 43, 10251011: 54,
    10251014: 54, 10251017: 48, 10251020: 41, 10251023: 45, 10251026: 81,
    10251029: 40, 10251032: 41, 10251035: 56, 10251038: 48, 10251044: 56,
    10251047: 51, 10251050: 44, 10251053: 34, 10251056: 43, 10251059: 43,
    10251062: 55, 10251065: 35, 10251068: 27, 10251071: 55, 10251074: 54,
    10251077: 62, 10251080: 59, 10251083: 70, 10251086: 82, 10251089: 55,
    10251092: 29, 10251095: 81, 10251104: 45, 10251107: 38, 10251110: 38,
    10251114: 60, 10251117: 52,
}

UAS_ASLI_B = {
    10231001: 50, 10231038: 40, 10251006: 42, 10251009: 20, 10251012: 56,
    10251015: 29, 10251018: 73, 10251021: 42, 10251024: 72, 10251027: 55,
    10251033: 76, 10251036: 32, 10251039: 37, 10251042: 42, 10251045: 20,
    10251048: 61, 10251051: 37, 10251054: 24, 10251057: 50, 10251060: 52,
    10251063: 52, 10251066: 39, 10251069: 12, 10251072: 63, 10251075: 22,
    10251078: 58, 10251081: 0, 10251084: 36, 10251087: 46, 10251090: 74,
    10251093: 12, 10251096: 0, 10251099: 54, 10251102: 32, 10251105: 16,
    10251108: 22, 10251112: 54, 10251115: 38,
}

# UAS FINAL setelah remedial UAS. Formula: UAS_final = MAX(UAS_asli, MIN(UAS_remedial, 65))
# KOSONG sampai remedial UAS dinilai (deadline submit Jumat, 19 Juni 2026 14.00 WITA).
# Isi hanya mhs yang skornya NAIK karena remedial; sisanya otomatis pakai UAS asli.
UAS_FINAL_OVERRIDE = {
    'A': {},
    'B': {},
}


# Sort order: pakai urutan README evaluasi dulu, lalu mahasiswa "extra" dari
# Excel (yang tidak ada di evaluasi) di-append di akhir.
ORDER_A = [
    10231081, 10251002, 10251005, 10251008, 10251011, 10251014, 10251017,
    10251020, 10251023, 10251026, 10251029, 10251032, 10251035, 10251038,
    10251044, 10251047, 10251050, 10251053, 10251056, 10251059, 10251062,
    10251065, 10251068, 10251071, 10251074, 10251086, 10251077, 10251080,
    10251083, 10251089, 10251092, 10251095, 10251104, 10251107, 10251110,
    10251114, 10251117,
]

ORDER_B = [
    10231001, 10231038, 10251006, 10251009, 10251012, 10251015, 10251018,
    10251021, 10251024, 10251027, 10251084, 10251033, 10251036, 10251039,
    10251042, 10251048, 10251051, 10251054, 10251057, 10251060, 10251063,
    10251066, 10251069, 10251072, 10251075, 10251078, 10251081, 10251087,
    10251090, 10251093, 10251099, 10251102, 10251105, 10251108, 10251112,
    10251115,
]


def find_evaluation_file(kelas, nim, username):
    folder = Path(f'Hasil-Evaluasi-Kelas-{kelas}')
    if not folder.exists():
        return None
    matches = list(folder.glob(f"Evaluasi_{nim}_*.md"))
    if matches:
        return matches[0].name
    if username:
        matches = list(folder.glob(f"Evaluasi_*_{username}.md"))
        if matches:
            return matches[0].name
    return None


def build_row_order(kelas, excel_data, order_list):
    """Return list NIM diurutkan ASCENDING berdasarkan NIM (semua mahasiswa,
    termasuk yang ada di Excel maupun di order_list)."""
    all_nims = set(excel_data.keys()) | set(order_list)
    return sorted(all_nims)


def generate_markdown(kelas, praktikum_map, username_map, order_list, excel_data,
                      uas_asli_map, output_path):
    full_order = build_row_order(kelas, excel_data, order_list)
    lines = []
    lines.append(f"# Nilai Final Praktikum Struktur Data (Kelas {kelas})")
    lines.append("")
    lines.append(f"> **Komposisi bobot:** Praktikum **{int(BOBOT['praktikum']*100)}%**, "
                 f"Kuis **{int(BOBOT['kuis']*100)}%**, UTS **{int(BOBOT['uts']*100)}%**, "
                 f"UAS **{int(BOBOT['uas']*100)}%**, Kehadiran **{int(BOBOT['hadir']*100)}%**.")
    lines.append("")
    lines.append("> **Kehadiran:** HADIR = 1, IZIN = 0.5, TANPA KETERANGAN = 0. Sel kosong pada "
                 "minggu yang **sudah berjalan** dianggap alpa (0); sel kosong pada minggu yang "
                 "**belum berjalan** di-skip. Skor = (sum / total minggu sudah berjalan) x 100.")
    lines.append("")
    lines.append("> **UTS & UAS:** ditampilkan nilai **final** (setelah remedial bila ada). "
                 "Formula remedial: `nilai_final = MAX(asli, MIN(remedial, 65))`.")
    lines.append("")
    lines.append("> **Catatan UAS:** kolom UAS = nilai **asli**. Remedial UAS (67 mhs) masih "
                 "berjalan — deadline submit **Jumat, 19 Juni 2026 14.00 WITA**. Nilai mhs peserta "
                 "remedial dapat naik (maks 65) dan **Nilai Akhir akan dihitung ulang** setelahnya.")
    lines.append("")
    lines.append("> **Konversi huruf** (batas bawah inklusif): A >= 86, AB 76-86, B 66-76, "
                 "BC 56-66, C 51-56, D 41-51, E < 41.")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## Rekapitulasi Nilai")
    lines.append("")
    lines.append("| No | NIM | Nama | Praktikum (30%) | Kuis (10%) | UTS (25%) | UAS (25%) | "
                 "Hadir (10%) | Nilai Akhir | Huruf | Detail |")
    lines.append("|---:|:---|:---|:---:|:---:|:---:|:---:|:---:|:---:|:---:|:---|")

    overrides = ATTENDANCE_OVERRIDE.get(kelas, {})
    uts_overrides = UTS_FINAL_OVERRIDE.get(kelas, {})
    kuis_overrides = KUIS_OVERRIDE.get(kelas, {})
    uas_overrides = UAS_FINAL_OVERRIDE.get(kelas, {})
    for idx, nim in enumerate(full_order, start=1):
        excel = excel_data.get(nim, {})
        nama_excel = excel.get('nama_excel', '?')
        nama = nama_excel if nama_excel else f"NIM {nim}"
        prak = praktikum_map.get(nim)
        kuis = kuis_overrides.get(nim, excel.get('kuis'))
        uts = uts_overrides.get(nim, excel.get('uts'))
        uas = uas_overrides.get(nim, uas_asli_map.get(nim))
        hadir = overrides.get(nim, excel.get('attendance'))
        username = username_map.get(nim, '')

        nilai = compute_nilai_akhir(prak, kuis, uts, uas, hadir)
        grade = letter_grade(nilai)

        eval_file = find_evaluation_file(kelas, nim, username)
        detail = f"[Lihat]({eval_file})" if eval_file else "-"

        lines.append(
            f"| {idx} | {nim} | **{nama}** | {fmt(prak)} | {fmt(kuis)} | {fmt(uts)} | {fmt(uas)} | "
            f"{fmt(hadir)} | **{fmt(nilai)}** | **{grade}** | {detail} |"
        )

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## Formula Perhitungan Final")
    lines.append("")
    lines.append("```")
    lines.append("Nilai Akhir = Praktikum x 0.30 + Kuis x 0.10 + UTS x 0.25 + UAS x 0.25 + Kehadiran x 0.10")
    lines.append("```")
    lines.append("")
    lines.append("UTS & UAS memakai nilai final pasca-remedial: `nilai_final = MAX(asli, MIN(remedial, 65))`.")
    lines.append("")
    lines.append("Remedial UAS masih berjalan (deadline 19 Juni 2026). Setelah dinilai, isi "
                 "`UAS_FINAL_OVERRIDE` di `compute_final_grades.py` lalu jalankan ulang script.")
    lines.append("")

    output_path.write_text('\n'.join(lines), encoding='utf-8')
    print(f"Written: {output_path}")


if __name__ == "__main__":
    data = parse_excel()
    generate_markdown('A', PRAKTIKUM_A, USERNAME_A, ORDER_A, data['A'], UAS_ASLI_A,
                      Path('Hasil-Evaluasi-Kelas-A/Nilai-Final-Kelas-A.md'))
    generate_markdown('B', PRAKTIKUM_B, USERNAME_B, ORDER_B, data['B'], UAS_ASLI_B,
                      Path('Hasil-Evaluasi-Kelas-B/Nilai-Final-Kelas-B.md'))
